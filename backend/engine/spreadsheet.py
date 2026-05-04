"""
SpreadsheetExtractionEngine: Excel/CSV parsing for result sheets.

Supports wide sheets with one student per row and long sheets with one
student-subject result per row.
"""

import csv
import io
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from models.schemas import StudentRecord, SubjectResult

GRADE_MAP = {
    "O": 10, "A+": 9, "A": 8, "B+": 7, "B": 6,
    "C": 5, "P": 4, "F": 0, "AB": 0, "NE": 0,
}

SUBJECT_CODE_RE = re.compile(r"\b(\d{2}[A-Z]{2,5}\d{2,3}|[A-Z]{2,5}\d{2,3})\b")

FIELD_ALIASES = {
    "sl_no": {"slno", "sno", "srno", "serialno", "serialnumber", "no"},
    "usn": {"usn", "universityseatnumber", "seatnumber", "register", "registerno", "regno"},
    "roll_no": {"rollno", "rollnumber", "roll", "classrollno"},
    "name": {"name", "studentname", "student"},
    "sgpa": {"sgpa", "semestergpa", "semestergp", "semestergradepointaverage"},
    "cgpa": {"cgpa", "cumulativegpa", "cumulativegradepointaverage"},
    "status": {"status", "result", "resultstatus", "passfail"},
    "total_credits": {"totalcredits", "totalcredit", "registeredcredits", "creditsregistered"},
    "credits_earned": {"creditsearned", "earnedcredits", "creditsobtained", "obtainedcredits"},
}


class SpreadsheetExtractionEngine:
    """Parses result spreadsheets into the same model used by PDF extraction."""

    def parse(self, file_bytes: bytes, filename: str) -> tuple[dict, list[StudentRecord]]:
        ext = Path(filename).suffix.lower()
        if ext in {".xlsx", ".xlsm"}:
            sheet_name, rows = self._load_xlsx(file_bytes)
        elif ext == ".csv":
            sheet_name, rows = "CSV Upload", self._load_csv(file_bytes)
        else:
            raise ValueError("Unsupported spreadsheet type. Upload .xlsx, .xlsm, or .csv.")

        rows = self._strip_empty_edges(rows)
        if not rows:
            raise ValueError("Spreadsheet is empty.")

        header_idx = self._find_header_row(rows)
        if header_idx is None:
            raise ValueError("Could not find a header row. Include columns such as USN, SGPA, and subject grades.")

        headers = [self._cell_text(v) for v in rows[header_idx]]
        data_rows = rows[header_idx + 1:]
        metadata = self._parse_metadata(rows[:header_idx], sheet_name)
        field_map = self._build_field_map(headers)

        if "usn" not in field_map and "name" not in field_map:
            raise ValueError("Spreadsheet must include at least a USN or student name column.")

        students = (
            self._parse_long_rows(headers, data_rows, field_map)
            if self._looks_like_long_format(headers, field_map)
            else self._parse_wide_rows(headers, data_rows, field_map)
        )

        if not students:
            raise ValueError("No student records found in the spreadsheet.")

        return metadata, students

    def _load_xlsx(self, file_bytes: bytes) -> tuple[str, list[list[Any]]]:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            best_sheet = None
            best_non_empty = 0
            best_rows: list[list[Any]] = []
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                non_empty = sum(1 for row in rows if any(self._cell_text(v) for v in row))
                if non_empty > best_non_empty:
                    best_sheet = sheet
                    best_non_empty = non_empty
                    best_rows = rows
            if best_sheet is None:
                raise ValueError("Workbook has no readable sheets.")
            return best_sheet.title, best_rows
        finally:
            workbook.close()

    def _load_csv(self, file_bytes: bytes) -> list[list[str]]:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")
        return [row for row in csv.reader(io.StringIO(text))]

    def _strip_empty_edges(self, rows: list[list[Any]]) -> list[list[Any]]:
        cleaned = [row for row in rows if any(self._cell_text(v) for v in row)]
        if not cleaned:
            return []
        max_cols = max(len(row) for row in cleaned)
        return [row + [None] * (max_cols - len(row)) for row in cleaned]

    def _find_header_row(self, rows: list[list[Any]]) -> Optional[int]:
        best_idx = None
        best_score = 0
        for idx, row in enumerate(rows[:25]):
            headers = [self._cell_text(v) for v in row]
            norms = [self._norm(h) for h in headers]
            score = 0
            for header, norm in zip(headers, norms):
                if any(norm in aliases for aliases in FIELD_ALIASES.values()):
                    score += 2
                if norm in FIELD_ALIASES["usn"]:
                    score += 3
                if self._subject_code_from_header(header):
                    score += 1
                if self._is_grade_header(norm):
                    score += 1
            if score > best_score:
                best_idx, best_score = idx, score
        return best_idx if best_score >= 3 else None

    def _parse_metadata(self, rows: list[list[Any]], sheet_name: str) -> dict:
        meta = {"college_name": "", "program": "", "semester": "", "exam_date": ""}
        for row in rows[:10]:
            line = " ".join(self._cell_text(v) for v in row if self._cell_text(v))
            upper = line.upper()
            if not line:
                continue
            if not meta["college_name"] and ("COLLEGE" in upper or "INSTITUTE" in upper):
                meta["college_name"] = line
            elif not meta["program"] and ("PROGRAM" in upper or "B.E" in upper or "BTECH" in upper):
                meta["program"] = line
            elif not meta["semester"] and ("SEM" in upper or "EXAM" in upper):
                meta["semester"] = line
            elif not meta["exam_date"] and "DATE" in upper:
                meta["exam_date"] = line
        if not meta["semester"]:
            meta["semester"] = sheet_name
        return meta

    def _build_field_map(self, headers: list[str]) -> dict[str, int]:
        field_map: dict[str, int] = {}
        for idx, header in enumerate(headers):
            norm = self._norm(header)
            for field, aliases in FIELD_ALIASES.items():
                if field not in field_map and norm in aliases:
                    field_map[field] = idx
        return field_map

    def _looks_like_long_format(self, headers: list[str], field_map: dict[str, int]) -> bool:
        if "usn" not in field_map:
            return False
        has_subject_code = any(self._is_subject_code_header(self._norm(h)) for h in headers)
        has_grade = any(self._is_grade_header(self._norm(h)) for h in headers)
        return has_subject_code and has_grade

    def _parse_long_rows(
        self,
        headers: list[str],
        rows: list[list[Any]],
        field_map: dict[str, int],
    ) -> list[StudentRecord]:
        subject_code_idx = self._first_header(headers, self._is_subject_code_header)
        grade_idx = self._first_header(headers, self._is_grade_header)
        students: OrderedDict[str, StudentRecord] = OrderedDict()

        for row in rows:
            usn = self._student_key(row, field_map)
            grade = self._parse_grade(self._value(row, grade_idx))
            if not usn or not grade:
                continue

            if usn not in students:
                seq = len(students) + 1
                students[usn] = self._record_from_row(row, field_map, seq, usn, [])

            code = self._cell_text(self._value(row, subject_code_idx)) or f"SUB{len(students[usn].subjects) + 1:02d}"
            students[usn].subjects.append(SubjectResult(code=code, grade=grade, grade_point=GRADE_MAP[grade]))
            self._refresh_record(students[usn], row, field_map)

        return list(students.values())

    def _parse_wide_rows(
        self,
        headers: list[str],
        rows: list[list[Any]],
        field_map: dict[str, int],
    ) -> list[StudentRecord]:
        meta_cols = set(field_map.values())
        subject_cols = self._subject_columns(headers, rows, meta_cols)
        students: list[StudentRecord] = []

        for row in rows:
            usn = self._student_key(row, field_map)
            if not usn and not self._value(row, field_map.get("sgpa")):
                continue

            subjects: list[SubjectResult] = []
            for pos, idx in enumerate(subject_cols, start=1):
                grade = self._parse_grade(self._value(row, idx))
                if not grade:
                    continue
                code = self._subject_code_from_header(headers[idx]) or self._clean_subject_name(headers[idx]) or f"SUB{pos:02d}"
                subjects.append(SubjectResult(code=code, grade=grade, grade_point=GRADE_MAP[grade]))

            if not usn and not subjects:
                continue

            students.append(self._record_from_row(row, field_map, len(students) + 1, usn, subjects))

        return students

    def _record_from_row(
        self,
        row: list[Any],
        field_map: dict[str, int],
        seq: int,
        usn: str,
        subjects: list[SubjectResult],
    ) -> StudentRecord:
        total = self._parse_int(self._value(row, field_map.get("total_credits")), 0)
        earned = self._parse_int(self._value(row, field_map.get("credits_earned")), total)
        status = self._parse_status(self._value(row, field_map.get("status")), subjects)
        return StudentRecord(
            sl_no=self._parse_int(self._value(row, field_map.get("sl_no")), seq),
            usn=usn or f"ROW{seq:03d}",
            roll_no=self._parse_optional_int(self._value(row, field_map.get("roll_no"))),
            name=self._cell_text(self._value(row, field_map.get("name"))),
            subjects=subjects,
            sgpa=self._parse_float(self._value(row, field_map.get("sgpa")), 0.0),
            cgpa=self._parse_float(self._value(row, field_map.get("cgpa")), 0.0),
            total_credits=total,
            credits_earned=earned,
            status=status,
        )

    def _refresh_record(self, record: StudentRecord, row: list[Any], field_map: dict[str, int]) -> None:
        if record.sgpa <= 0:
            record.sgpa = self._parse_float(self._value(row, field_map.get("sgpa")), record.sgpa)
        if record.cgpa <= 0:
            record.cgpa = self._parse_float(self._value(row, field_map.get("cgpa")), record.cgpa)
        if not record.name:
            record.name = self._cell_text(self._value(row, field_map.get("name")))
        record.status = self._parse_status(self._value(row, field_map.get("status")), record.subjects)

    def _subject_columns(self, headers: list[str], rows: list[list[Any]], meta_cols: set[int]) -> list[int]:
        cols = []
        for idx, header in enumerate(headers):
            if idx in meta_cols or not header:
                continue
            norm = self._norm(header)
            if self._is_subject_code_header(norm) or self._is_grade_header(norm):
                continue
            if self._subject_code_from_header(header) or self._clean_subject_name(header):
                sample = [self._parse_grade(self._value(row, idx)) for row in rows[:40]]
                if any(sample):
                    cols.append(idx)
        return cols

    def _student_key(self, row: list[Any], field_map: dict[str, int]) -> str:
        usn = self._cell_text(self._value(row, field_map.get("usn"))).upper()
        if usn and self._norm(usn) not in FIELD_ALIASES["usn"]:
            return usn
        name = self._cell_text(self._value(row, field_map.get("name")))
        roll = self._cell_text(self._value(row, field_map.get("roll_no")))
        return f"{roll}-{name}".strip("-").upper()

    def _parse_status(self, value: Any, subjects: list[SubjectResult]) -> str:
        text = self._cell_text(value).upper()
        if text in {"P", "PASS", "PASSED"}:
            return "P"
        if text in {"F", "FAIL", "FAILED"}:
            return "F"
        if text in {"NP", "NE", "AB"}:
            return text
        if any(s.grade == "F" for s in subjects):
            return "F"
        if any(s.grade == "AB" for s in subjects):
            return "AB"
        if any(s.grade == "NE" for s in subjects):
            return "NE"
        return "P"

    def _parse_grade(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            points = int(value)
            return {10: "O", 9: "A+", 8: "A", 7: "B+", 6: "B", 5: "C", 4: "P", 0: "F"}.get(points)

        text = self._cell_text(value).upper().replace("PLUS", "+")
        text = re.sub(r"\s+", " ", text).strip()
        compact = text.replace(" ", "")
        if compact in GRADE_MAP:
            return compact
        if compact in {"PASS", "PASSED"}:
            return "P"
        if compact in {"FAIL", "FAILED"}:
            return "F"
        if compact.isdigit():
            return self._parse_grade(int(compact))

        match = re.search(r"\b(A\+|B\+|AB|NE|O|A|B|C|P|F)\b", text)
        return match.group(1) if match else None

    def _subject_code_from_header(self, header: str) -> str:
        match = SUBJECT_CODE_RE.search(self._cell_text(header).upper())
        return match.group(1) if match else ""

    def _clean_subject_name(self, header: str) -> str:
        text = re.sub(r"\b(grade|gp|grade point|marks|credit|credits|earned)\b", "", self._cell_text(header), flags=re.I)
        text = re.sub(r"\s+", " ", text).strip(" :-_")
        return "" if self._norm(text) in {"", "subject", "sub"} else text[:30]

    def _first_header(self, headers: list[str], predicate) -> Optional[int]:
        for idx, header in enumerate(headers):
            if predicate(self._norm(header)):
                return idx
        return None

    def _is_subject_code_header(self, norm: str) -> bool:
        return norm in {"subjectcode", "subcode", "coursecode", "papercode", "subject"}

    def _is_grade_header(self, norm: str) -> bool:
        return norm in {"grade", "grades", "subjectgrade", "resultgrade", "gradeobtained"}

    def _value(self, row: list[Any], idx: Optional[int]) -> Any:
        return row[idx] if idx is not None and idx < len(row) else None

    def _cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _norm(self, value: str) -> str:
        return re.sub(r"[^a-z0-9+]", "", self._cell_text(value).lower())

    def _parse_float(self, value: Any, default: float) -> float:
        if value is None or value == "":
            return default
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        match = re.search(r"-?\d+(?:\.\d+)?", self._cell_text(value).replace(",", ""))
        return float(match.group(0)) if match else default

    def _parse_int(self, value: Any, default: int) -> int:
        return int(round(self._parse_float(value, float(default))))

    def _parse_optional_int(self, value: Any) -> Optional[int]:
        text = self._cell_text(value)
        return self._parse_int(text, 0) if text else None
